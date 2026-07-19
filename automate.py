#!/usr/bin/env python3
"""
杉树公益基金会 - 面试日程表 & 评估表生成 + 结果汇总
"""
import json
import logging
from datetime import datetime, timedelta
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

# ── 样式定义 ─────────────────────────────────────────
HEADER_FILL  = PatternFill("solid", fgColor="4472C4")
HEADER_FONT  = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT   = Font(name="微软雅黑", bold=True, size=14)
BODY_FONT    = Font(name="微软雅黑", size=10)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_WRAP    = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
LIGHT_GREEN  = PatternFill("solid", fgColor="E2EFDA")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFF2CC")
LIGHT_RED    = PatternFill("solid", fgColor="FCE4EC")

# ── 数据加载（从 CSV 读取） ──────────────────────────────
CSV_PATH = "candidates.csv"
INTERVIEWERS = [
    {"name": "李老师", "role": "项目主管"},
    {"name": "王老师", "role": "教学组长"},
    {"name": "张老师", "role": "心理辅导师"},
]

def load_candidates(csv_path: str = CSV_PATH) -> list[dict]:
    """从 CSV 读取候选人数据。CSV 格式：姓名,手机号,科目,支教学校,可面试时间"""
    p = Path(csv_path)
    if not p.exists():
        logging.warning(f"⚠️ 未找到 {csv_path}，使用示例数据")
        return [
            {"name": "张三", "phone": "13800001001", "subject": "数学", "school": "云南昭通某小学", "available": "全天"},
            {"name": "李四", "phone": "13800001002", "subject": "语文", "school": "四川凉山某小学", "available": "上午"},
            {"name": "王五", "phone": "13800001003", "subject": "英语", "school": "贵州毕节某小学", "available": "下午"},
            {"name": "赵六", "phone": "13800001004", "subject": "数学", "school": "甘肃会宁某小学", "available": "全天"},
            {"name": "陈七", "phone": "13800001005", "subject": "科学", "school": "湖南湘西某小学", "available": "上午"},
        ]

    import csv
    candidates = []
    with p.open(newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            candidates.append({
                "name": row.get("姓名", row.get("name", "")).strip(),
                "phone": row.get("手机号", row.get("phone", "")).strip(),
                "subject": row.get("科目", row.get("subject", "")).strip(),
                "school": row.get("支教学校", row.get("school", "")).strip(),
                "available": row.get("可面试时间", row.get("available", "")).strip(),
            })
    logging.info(f"📄 已从 {csv_path} 读取 {len(candidates)} 位候选人")
    return candidates

# ── 1. 面试日程表 ────────────────────────────────────
def gen_schedule(candidates: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "面试日程表"

    # 标题
    ws.merge_cells("A1:G1")
    ws["A1"] = "上海杉树公益基金会 · 支教老师面试日程表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ["序号", "候选人", "科目", "支教学校", "面试时间", "面试形式", "面试官"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 28

    # 数据行
    base_time = datetime(2026, 8, 10, 9, 0)  # 8月10日9:00开始
    for i, c in enumerate(candidates, 1):
        row = i + 3
        t = base_time + timedelta(minutes=30 * (i - 1))
        time_str = t.strftime("%m/%d %H:%M")
        form = "线上面试" if i % 2 == 1 else "线下面试"
        interviewer = INTERVIEWERS[(i - 1) % len(INTERVIEWERS)]["name"]

        values = [i, c["name"], c["subject"], c["school"], time_str, form, interviewer]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            # 交替行颜色
            if i % 2 == 0:
                cell.fill = LIGHT_GREEN
        ws.row_dimensions[row].height = 24

    # 列宽
    widths = [8, 10, 8, 26, 16, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = "面试日程表.xlsx"
    wb.save(path)
    logging.info(f"✅ 面试日程表已生成: {path}")
    return path


# ── 2. 面试评估表（每人一份） ──────────────────────────
def gen_evaluation_sheets(candidates: list[dict]):
    paths = []
    for c in candidates:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "面试评估表"

        # 标题
        ws.merge_cells("A1:F1")
        ws["A1"] = f"杉树公益 · 支教面试评估表 —— {c['name']}"
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = CENTER
        ws.row_dimensions[1].height = 36

        # 基本信息
        info = [["候选人", c["name"], "科目", c["subject"], "", ""],
                ["支教学校", c["school"], "联系电话", c["phone"], "", ""],
                ["面试官", "_______________", "面试日期", "____年__月__日", "", ""]]
        for ri, row_data in enumerate(info, 2):
            for ci, v in enumerate(row_data, 1):
                cell = ws.cell(row=ri, column=ci, value=v)
                cell.font = BODY_FONT
                cell.alignment = LEFT_WRAP
                cell.border = THIN_BORDER
                if ci in (1, 3):
                    cell.fill = PatternFill("solid", fgColor="D9E2F3")

        # 评分项
        score_row = 6
        eval_headers = ["评分项", "权重", "评分(1-5)", "加权得分", "评分说明"]
        for col, h in enumerate(eval_headers, 1):
            cell = ws.cell(row=score_row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        items = [
            ("教育理念与使命感", "20%", "对教育公益的理解与认同"),
            ("教学能力与表达", "25%", "语言表达、逻辑思维"),
            ("应变与抗压能力", "15%", "面对困难的应对方式"),
            ("团队协作与沟通", "15%", "与人合作的经验与意愿"),
            ("责任心与耐心", "15%", "对待孩子的态度"),
            ("特长与加分项", "10%", "音乐/体育/美术等特长"),
        ]
        for i, (item, weight, note) in enumerate(items):
            row = score_row + 1 + i
            ws.cell(row=row, column=1, value=item).font = BODY_FONT
            ws.cell(row=row, column=1).border = THIN_BORDER
            ws.cell(row=row, column=1).alignment = LEFT_WRAP
            ws.cell(row=row, column=2, value=weight).font = BODY_FONT
            ws.cell(row=row, column=2).alignment = CENTER
            ws.cell(row=row, column=2).border = THIN_BORDER
            ws.cell(row=row, column=3).border = THIN_BORDER  # 评分格
            ws.cell(row=row, column=3).alignment = CENTER
            ws.cell(row=row, column=4).border = THIN_BORDER  # 加权得分
            ws.cell(row=row, column=4).alignment = CENTER
            ws.cell(row=row, column=5, value=note).font = Font(name="微软雅黑", size=9, color="888888")
            ws.cell(row=row, column=5).border = THIN_BORDER
            ws.cell(row=row, column=5).alignment = LEFT_WRAP

        # 总分 & 建议
        total_row = score_row + len(items) + 2
        ws.cell(row=total_row, column=1, value="总分").font = Font(name="微软雅黑", bold=True, size=11)
        ws.cell(row=total_row, column=1).border = THIN_BORDER
        ws.cell(row=total_row, column=2).border = THIN_BORDER
        ws.cell(row=total_row, column=1).fill = LIGHT_YELLOW
        ws.cell(row=total_row, column=2).fill = LIGHT_YELLOW
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)

        result_row = total_row + 1
        ws.cell(row=result_row, column=1, value="面试建议").font = Font(name="微软雅黑", bold=True, size=11)
        ws.cell(row=result_row, column=1).alignment = LEFT_WRAP
        ws.cell(row=result_row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=result_row, start_column=1, end_row=result_row, end_column=5)
        ws.cell(row=result_row, column=5).border = THIN_BORDER
        ws.row_dimensions[result_row].height = 60

        # 列宽
        for i, w in enumerate([20, 10, 12, 12, 30], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        safe = c["name"]
        path = f"面试评估表_{safe}.xlsx"
        wb.save(path)
        paths.append(path)
        logging.info(f"✅ 评估表已生成: {path}")

    return paths


# ── 3. 面试结果汇总表 ────────────────────────────────
def gen_summary(candidates: list[dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "面试结果汇总"

    # 标题
    ws.merge_cells("A1:H1")
    ws["A1"] = "上海杉树公益基金会 · 面试结果汇总表"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ["序号", "候选人", "科目", "支教学校", "综合评分", "结果", "状态", "备注"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[3].height = 28

    # 模拟评分结果（实际使用时从评估表读取）
    import random
    random.seed(42)
    results = []
    for i, c in enumerate(candidates):
        score = round(random.uniform(3.0, 5.0), 1)
        passed = score >= 3.5
        status = "✅ 岗前培训" if passed else "❌ 未通过"
        note = "" if passed else "进入人才库"
        results.append((i + 1, c["name"], c["subject"], c["school"], score, "通过" if passed else "未通过", status, note))

    for ri, (seq, name, subj, sch, score, result, st, note) in enumerate(results):
        row = ri + 4
        vals = [seq, name, subj, sch, score, result, st, note]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.font = BODY_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER
            if result == "通过":
                cell.fill = LIGHT_GREEN if col not in (7,) else PatternFill("solid", fgColor="C6EFCE")
            else:
                cell.fill = LIGHT_RED
        ws.row_dimensions[row].height = 24

    # 统计行
    stats_row = len(results) + 5
    passed_count = sum(1 for r in results if r[5] == "通过")
    ws.cell(row=stats_row, column=1, value=f"统计：共 {len(results)} 人，通过 {passed_count} 人，未通过 {len(results) - passed_count} 人")
    ws.cell(row=stats_row, column=1).font = Font(name="微软雅黑", bold=True, size=11)
    ws.merge_cells(start_row=stats_row, start_column=1, end_row=stats_row, end_column=8)

    # 列宽
    for i, w in enumerate([8, 10, 8, 26, 12, 10, 16, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    path = "面试结果汇总表.xlsx"
    wb.save(path)
    logging.info(f"✅ 结果汇总表已生成: {path}")
    return path, results


# ── 主入口（配合 GitHub Actions 的 --retry 参数） ─────
def main():
    MAX_RETRIES = 0
    import sys
    if "--retry" in sys.argv:
        idx = sys.argv.index("--retry")
        MAX_RETRIES = int(sys.argv[idx + 1]) if idx + 1 < len(sys.argv) else 1

    logging.info(f"杉树公益 · 面试流程自动化开始 (retry={MAX_RETRIES})")

    try:
        candidates = load_candidates()
        schedule_path = gen_schedule(candidates)
        eval_paths = gen_evaluation_sheets(candidates)
        summary_path, results = gen_summary(candidates)

        # 生成汇总文件清单作为 artifact
        summary_data = {
            "生成时间": datetime.now().isoformat(),
            "面试日程表": schedule_path,
            "评估表数量": len(eval_paths),
            "评估表列表": eval_paths,
            "结果汇总表": summary_path,
            "面试人数": len(results),
            "通过人数": sum(1 for r in results if r[5] == "通过"),
            "未通过人数": sum(1 for r in results if r[5] == "未通过"),
        }
        Path("output.json").write_text(json.dumps(summary_data, ensure_ascii=False, indent=2))
        total_files = 1 + len(eval_paths) + 1  # schedule + evals + summary
        logging.info(f"✅ 全部完成！生成文件：{total_files} 份")
        print(json.dumps(summary_data, ensure_ascii=False))

    except Exception as e:
        logging.error(f"❌ 自动化失败: {e}")
        Path("error.log").write_text(f"{datetime.now()}: {e}")
        raise


if __name__ == "__main__":
    main()
